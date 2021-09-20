from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from openpyxl import load_workbook
from pprint import pprint
import statistics
import os
from tkinter import *


def kit():
    def getSampleReactiveData():
        from openpyxl import load_workbook
        wb = load_workbook(filename='prelim_report.xlsx')
        inpt_sht = wb.active
        react_data = []
        for i in range(3, len(inpt_sht['K']) + 1):
            temp = "K" + str(i)
            react_data.append(inpt_sht[temp].value)
        return react_data

    wb_imported_input = load_workbook(filename='Input.xlsx')
    inputt = wb_imported_input["Input"]
    # Input sheet opened

    flag = 1
    list_ids = inputt['A']
    for x in range(2, len(list_ids)):
        if (list_ids[x].value == list_ids[x - 1].value):
            flag = flag + 1
            present = inputt['A' + str(x + 1)].value
            if present != None:
                inputt['A' + str(x + 1)].value = float(present) + flag * 0.1
        else:
            flag = 1

    inpt_workbook = Workbook()
    # Kit excel file created
    sym = ['/', '*', '[', ']', ':', '?', '-']
    kit_names = []
    for i in range(2, len(inputt['B'])):
        bb = str(inputt['B' + str(i)].value)
        for k in bb:
            if k in sym:
                bb = bb.replace(k, " ")
        kit_names.append(bb.upper())
    list_set = set(kit_names)
    unique_names = (list(list_set))
    if None in unique_names:
        unique_names.remove(None)
    unique_names.sort()
    # unique kit names taken

    fmt_bold = Font(bold=True)

    for name in unique_names:
        a = "A"
        inpt_worksheet = inpt_workbook.create_sheet(name)
        another_other_list = []
        for p in range(1, len(inputt[1]) + 1):
            another_other_list.append(inputt[a + str(1)].value)
            a = chr(ord(a) + 1)
        inpt_worksheet.append(another_other_list)
        a = "A"
        for p in range(1, len(inputt[1]) + 1):
            inpt_worksheet[a + str(1)].font = fmt_bold
            a = chr(ord(a) + 1)
    # All the different sheets with kits created

    alpha = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

    for name in unique_names:
        ws1 = inpt_workbook[name]
        for i in range(2, len(inputt['B']) + 1):
            if (inputt['B' + str(i)].value == name):
                another_list = []
                for al in alpha:
                    try:
                        paras = inputt[al + str(i)].value
                        paras = float(paras)
                        another_list.append(paras)
                    except:
                        another_list.append(inputt[al + str(i)].value)
                ws1.append(another_list)
        # All the values for every kit entered from input table

        inpt_workbook.save("Kitwise.xlsx")

    inpt_workbook = load_workbook(filename='Kitwise.xlsx')
    for name in unique_names:
        ws1 = inpt_workbook[name]
        flag = 0
        rec_stat = getSampleReactiveData()

        dee = 'C'
        sample_values = [ws1['D'], ws1['E'], ws1['F'], ws1['G'], ws1['H'], ws1['I'], ws1['J'], ws1['K']]
        leng = ws1.max_row

        for j in range(0, 8):
            if (ws1.max_row > 5):

                def put():
                    enter_screen.destroy()

                enter_screen = Tk()
                enter_screen.title("EQAS")
                enter_screen.geometry("400x400")
                Label(enter_screen,
                      text="[IMP] Please check the values for " + name + " kit \nbefore entering upper and lower limit values.").pack()
                Label(enter_screen, text="").pack()
                Label(enter_screen, text="Enter the lower limit for sample " + str(j + 1) + " " + name).pack()
                upper = DoubleVar()
                lower = DoubleVar()
                upper_entry = Entry(enter_screen, textvariable=upper)
                upper_entry.pack()
                Label(enter_screen, text="").pack()
                Label(enter_screen, text="Enter the upper limit:").pack()
                lower_entry = Entry(enter_screen, textvariable=lower)
                lower_entry.pack()
                Button(enter_screen, text="Enter", width=10, height=1, command=put).pack(pady=10)

                enter_screen.mainloop()
                low = lower.get()
                upp = upper.get()

            else:
                flag = 1
            # Conditional inclusion for kits with less than 5 entries done
            x = 0
            dee = chr(ord(dee) + 1)
            include = []
            if leng > 5:
                for i in range(1, leng):
                    # print(sample_values[j][i].value)
                    if (sample_values[j][i].value is not None):
                        if (sample_values[j][i].value == 'NR' or sample_values[j][i].value == 'R'):
                            if (rec_stat[j] == 'R' and sample_values[j][i].value == 'NR'):
                                temp = dee + str(i + 1)
                                ws1[temp].fill = PatternFill(fill_type='solid', start_color='ef9a9a',
                                                             end_color='ef9a9a')
                            elif (rec_stat[j] == 'NR' and sample_values[j][i].value == 'R'):
                                temp = dee + str(i + 1)
                                ws1[temp].fill = PatternFill(fill_type='solid', start_color='ef9a9a',
                                                             end_color='ef9a9a')
                        else:
                            if ((float(sample_values[j][i].value) < low or float(
                                    sample_values[j][i].value) > upp) and flag == 0):
                                temp = dee + str(i + 1)
                                ws1[temp].fill = PatternFill(fill_type='solid', start_color='64b5f6',
                                                             end_color='64b5f6')

                            elif (rec_stat[j] == 'R' and float(sample_values[j][i].value) < 1):
                                temp = dee + str(i + 1)
                                ws1[temp].fill = PatternFill(fill_type='solid', start_color='ef9a9a',
                                                             end_color='ef9a9a')

                            elif (rec_stat[j] == 'NR' and float(sample_values[j][i].value) > 1):
                                temp = dee + str(i + 1)
                                ws1[temp].fill = PatternFill(fill_type='solid', start_color='ef9a9a',
                                                             end_color='ef9a9a')

                            else:
                                include.append(float(sample_values[j][i].value))
                    else:
                        continue
            # Colouring done and kits with less than 5 entries will exit the loop from here and not be included for statistics

            if (flag == 1):
                goon = 0
            else:
                if (len(include) == 0):
                    continue
                avg = statistics.mean(include)
                med = statistics.median(include)
                ws1.cell(row=leng + 5, column=3).value = "MEAN"
                # wks.update_cell(leng+5,3,"MEAN")
                ws1.cell(row=leng + 6, column=3).value = "MEDIAN"
                # wks.update_cell(leng+6,3,"MEDIAN")
                ws1.cell(row=leng + 11, column=3).value = "Number_of_samples_for_evaluation"
                # wks.update_cell(leng+11,3,"Number_of_samples_for_evaluation")
                ws1.cell(row=leng + 11, column=j + 4).value = len(include)
                # wks.update_cell(leng+11,j+4,len(include))
                ws1.cell(row=leng + 5, column=j + 4).value = avg
                # wks.update_cell(leng+5,j+4,avg)
                ws1.cell(row=leng + 6, column=j + 4).value = med
                # wks.update_cell(leng+6,j+4,med)
                if (len(include) <= 1):
                    continue
                standarddev = statistics.stdev(include)
                cv = standarddev / avg * 100
                ws1.cell(row=leng + 7, column=3).value = "STANDARD DEV"
                # wks.update_cell(leng+7,3,"STANDARD DEV")
                ws1.cell(row=leng + 8, column=3).value = "CV"
                # wks.update_cell(leng+8,3,"CV")
                ws1.cell(row=leng + 9, column=3).value = "UPPER LIMIT"
                # wks.update_cell(leng+9,3,"UPPER LIMIT")
                ws1.cell(row=leng + 10, column=3).value = "LOWER LIMIT"
                # wks.update_cell(leng+10,3,"LOWER LIMIT")
                ws1.cell(row=leng + 7, column=j + 4).value = standarddev
                # wks.update_cell(leng+7,j+4,standarddev)
                ws1.cell(row=leng + 8, column=j + 4).value = cv
                # wks.update_cell(leng+8,j+4,cv)
                ws1.cell(row=leng + 9, column=j + 4).value = avg + 2 * standarddev
                # wks.update_cell(leng+9,j+4,avg+2*standarddev)
                ws1.cell(row=leng + 10, column=j + 4).value = avg - 2 * standarddev
                # wks.update_cell(leng+10,j+4,avg-2*standarddev)
                try:
                    for i in range(1, leng):
                        if ((low <= float(sample_values[j][i].value) <= upp) and (
                                rec_stat[j] == 'R' and float(sample_values[j][i].value) >= 1 or (
                                rec_stat[j] == 'NR' and float(sample_values[j][i].value) <= 1)) and (
                                float(sample_values[j][i].value) < avg - 2 * standarddev or float(
                                sample_values[j][i].value) > avg + 2 * standarddev)):
                            temp = dee + str(i + 1)
                            ws1[temp].fill = PatternFill(fill_type='solid', start_color='32CD32', end_color='32CD32')
                except:
                    pass

    inpt_workbook.remove(inpt_workbook['Sheet'])
    # removed the extra sheet created at beginning of object creation

    os.remove("Kitwise.xlsx")
    # temporary removed

    inpt_workbook.save("Kitwise.xlsx")
    # file saved
