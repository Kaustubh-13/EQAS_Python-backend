from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from openpyxl import load_workbook
from pprint import pprint
import os
from tkinter import messagebox


def inp():
    def getSampleReactiveData():
        from openpyxl import load_workbook
        wb = load_workbook(filename='prelim_report.xlsx')
        inpt_sht = wb.active
        react_data = []
        for i in range(3, len(inpt_sht['K']) + 1):
            temp = "K" + str(i)
            react_data.append(inpt_sht[temp].value)
        return react_data

    wb_imported_input = load_workbook(filename='Input.xlsx')
    input_sheet = wb_imported_input["Input"]
    # Input sheet opened

    try:
        react_data = getSampleReactiveData()
    except:
        MsgBox = messagebox.showinfo('file not found',
                                     'Please check the file \"prelim_report.xlsx\" and try running the code again.')
        exit()
    # sample's data is collected :: R/NR

    dee = 'C'
    for i in range(1, 9):
        dee = chr(ord(dee) + 1)
        col = input_sheet[dee]
        for j in range(1, len(col)):
            if input_sheet[dee + str(j + 1)].value != None and str(input_sheet[dee + str(j + 1)].value) < "A":
                ffe = float(input_sheet[dee + str(j + 1)].value)
                if ffe < 1 and react_data[i-1] == "R":
                    input_sheet[dee + str(j + 1)].fill = PatternFill(fill_type='solid', start_color='ef9a9a',
                                                                     end_color='ef9a9a')
                if ffe > 1 and react_data[i-1] == "NR":
                    input_sheet[dee + str(j + 1)].fill = PatternFill(fill_type='solid', start_color='ef9a9a',
                                                                     end_color='ef9a9a')
            elif input_sheet[dee + str(j + 1)].value == "NR" and react_data[i-1] == "R":
                input_sheet[dee + str(j + 1)].fill = PatternFill(fill_type='solid', start_color='ef9a9a', end_color='ef9a9a')
            elif input_sheet[dee + str(j + 1)].value == "R" and react_data[i-1] == "NR":
                input_sheet[dee + str(j + 1)].fill = PatternFill(fill_type='solid', start_color='ef9a9a', end_color='ef9a9a')
            else:
                continue
    # Colours all the aberrant values
    try:
        os.remove("Input.xlsx")
    except:
        pass

    try:
        wb_imported_input.save('Input.xlsx')
    except:
        msgBox = messagebox.showerror('file exists', 'Kindly delete/move the existing file and run the code again.')
    # Check if the file exists and throw an exception
