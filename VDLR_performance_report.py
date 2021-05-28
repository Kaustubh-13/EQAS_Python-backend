from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from openpyxl import load_workbook
from pprint import pprint
import os
from copy import copy

wb_kitwise = load_workbook(filename='Kitwise.xlsx')
# kit-wise_sheet_opened

wb_input_temp = load_workbook(filename='Input.xlsx')
wb_temp_sheet = wb_input_temp["Input"]
# temp input sheet

wb_perfotemp = load_workbook(filename='performance_report_template.xlsx')
perfotemp_hiv = wb_perfotemp["VDLR "]
# HIV performance template loaded

kit_names = list(wb_kitwise.sheetnames)
# kit_names taken

try:
    os.makedirs("VDLR")
except:
    pass

for sheet in kit_names:
    wb_active = wb_kitwise[sheet]
    temp_lab_ids = wb_active['A']
    lab_ids = []
    real_labs = []

    for cont in range(1, len(temp_lab_ids)):
        if (temp_lab_ids[cont].value is not None):
            lab_ids.append(temp_lab_ids[cont])

    leng = len(lab_ids)
    part = []
    for i in range(0, leng):
        part.append(int(lab_ids[i].value))
    means = wb_active["D" + str(leng + 6) + ":" + "K" + str(leng + 6)]
    stdev = wb_active["D" + str(leng + 8) + ":" + "K" + str(leng + 8)]
    cv = wb_active["D" + str(leng + 9) + ":" + "K" + str(leng + 9)]
    statees = wb_active["D" + str(leng + 12) + ":" + "K" + str(leng + 12)]
    pos = 1

    for id_cells in lab_ids:  # cell loop for lab ids
        pos = pos + 1
        if (id_cells.value == "ID" or id_cells.value is None):
            continue

        perfotemp_hiv['B4'].value = "Lab_name_goes_here"
        perfotemp_hiv['B4'].font = Font(bold=True)
        actual = float(id_cells.value)
        actual = actual * 10
        run = actual % 10
        lab_no = actual / 10

        if (run == 0):
            for z in range(10, 18):
                perfotemp_hiv['B' + str(z)] = 1
        else:
            for z in range(10, 18):
                perfotemp_hiv['B' + str(z)] = run

        perfotemp_hiv['B5'].value = lab_no
        perfotemp_hiv['B6'].value = sheet
        result = wb_active["D" + str(pos) + ":" + "K" + str(pos)]
        list_set = set(part)
        unique_names = (list(list_set))
        participants = part
        for j in range(10, 18):
            perfotemp_hiv["C" + str(j)].value = result[0][j - 10].value
            fmt = copy(result[0][j - 10].fill)
            perfotemp_hiv["D" + str(j)].fill = fmt
            # try:
            if (result[0][j - 10].value is None):
                perfotemp_hiv["L" + str(j)].value = ''  # if(result[0][j-10].value is not None):
            else:
                # if(float(result[0][j-10].value)):
                if (result[0][j - 10].fill.start_color.index == "00000000"):
                    perfotemp_hiv["L" + str(j)].value = 'Gud boi'
                else:
                    perfotemp_hiv["L" + str(j)].value = 'Bad boi'
            # except:
            #    pass
            # if(result[0][j-10].value is None):
            #    perfotemp_hiv["L"+str(j)].value=''
            perfotemp_hiv["H" + str(j)].value = means[0][j - 10].value
            perfotemp_hiv["I" + str(j)].value = stdev[0][j - 10].value
            perfotemp_hiv["J" + str(j)].value = cv[0][j - 10].value
            perfotemp_hiv["E" + str(j)].value = len(unique_names)
            perfotemp_hiv["F" + str(j)].value = len(part)
            perfotemp_hiv["G" + str(j)].value = statees[0][j - 10].value
            perfotemp_hiv["K" + str(j)].value = statees[0][j - 10].value
        try:
            os.makedirs("D:\\VDLR\\Lab_ID_" + str(int(lab_no)))
        except:
            pass
        if (run == 0):
            temp_run = 1
        else:
            temp_run = run
        wb_perfotemp.save(
            "D:\\VDLR\\Lab_ID_" + str(int(lab_no)) + "\\VDLR_" + str(int(lab_no)) + "_Run_" + str(int(temp_run)) + ".xlsx")
